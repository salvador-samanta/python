from openpyxl import load_workbook
import pathlib
import sqlite3

# pedimos la ruta del directorio donde tenemos guardado este .py
ruta = str(pathlib.Path(__file__).parent.absolute())

# asignamos la ruta del directorio y el acrhivo excel
archivo = ruta + "/movies.xlsx"

# creamos una lista con los nombres de las hojas.
lista = ["actor","director","idioma","pais","peliculas","tema"]

# hacemos la carga del excel.
libro = load_workbook(archivo, read_only=True)

# activamos las hojas. '0' es la posicion de la 1er hoja 
act_actor = libro[lista[0]]
act_director = libro[lista[1]]
act_idioma = libro[lista[2]]
act_pais = libro[lista[3]]
act_peliculas = libro[lista[4]]
act_tema = libro[lista[5]]

# creamos las listas de cada hoja para traer los datos
lista_actores = []
lista_director = []
lista_idioma = []
lista_pais = []
lista_peliculas = []
lista_tematica = []

lista_temporal = []

# lista_actores
# puede que de un error 'too many values to unpack (expect 4). 
# Hay que eliminar las columnas o filas vacias puede que hayamos marcado 
# algo que no se ve. Alguna celda que copiamos y borramos, por mas que
# veamos vacias, estan activas.
for a, b, c, d in act_actor.iter_rows(2):
    lista_temporal.append(a.value)
    lista_temporal.append(b.value)
    lista_temporal.append(c.value)
    lista_temporal.append(d.value)

    lista_actores.append(lista_temporal)
    
    lista_temporal = []

print(lista_actores)
input("\nCarga lista_actores finalizada. presione <ENTER>.")
print()
    
# lista_director
for a, b, c, d in act_director.iter_rows(2):
    lista_temporal.append(a.value)
    lista_temporal.append(b.value)
    lista_temporal.append(c.value)
    lista_temporal.append(d.value)

    lista_director.append(lista_temporal)
    
    lista_temporal = []

print(lista_director)
input("\nCarga lista_director finalizada. presione <ENTER>.")
print()

# lista_idioma
for a, b in act_idioma.iter_rows(2):
    lista_temporal.append(a.value)
    lista_temporal.append(b.value)

    lista_idioma.append(lista_temporal)
    
    lista_temporal = []

print(lista_idioma)

input("\nCarga lista_idioma finalizada. presione <ENTER>.")
print()

# lista_pais
for a, b in act_pais.iter_rows(2):
    lista_temporal.append(a.value)
    lista_temporal.append(b.value)

    lista_pais.append(lista_temporal)
    
    lista_temporal = []

print(lista_pais)

input("\nCarga lista_pais finalizada. presione <ENTER>.")
print()

# lista_peliculas
for a, b, c, d, e, f, g, h, i, j in act_peliculas.iter_rows(2):
    lista_temporal.append(a.value)
    lista_temporal.append(b.value)
    lista_temporal.append(c.value)
    lista_temporal.append(d.value)
    lista_temporal.append(e.value)
    lista_temporal.append(f.value)
    lista_temporal.append(g.value)
    lista_temporal.append(h.value)
    lista_temporal.append(float(i.value))
    lista_temporal.append(j.value)

    lista_peliculas.append(lista_temporal)
    
    lista_temporal = []

print(lista_peliculas)
input("\nCarga lista_peliculas finalizada. presione <ENTER>.")
print()

# lista_tematica
for a, b in act_tema.iter_rows(2):
    lista_temporal.append(a.value)
    lista_temporal.append(b.value)

    lista_tematica.append(lista_temporal)
    
    lista_temporal = []

print(lista_tematica)
input("\nCarga lista_tematica finalizada. presione <ENTER>.")
print()

# Carga SQLITE 3

# generamos la conexion a 'movies.db'. Usamos la variable 
# 'ruta' para indicarle donde se encuentra el archivo.
conexion = sqlite3.connect(ruta + "/movies.db")
cur = conexion.cursor()

print("Se establecio conexion.\n")
# insertamos informacion en los camposdatetime A combination of a date and a time. 
for i in lista_pais:
    inserta="INSERT INTO pais values (?, ?);"
    cur.execute(inserta, i)
    conexion.commit()

input("\nSe cargo tabla 'pais' en 'movies.db'.")

for i in lista_director:
    inserta="INSERT INTO director values (?, ?, ?, ?);"
    cur.execute(inserta, i)
    conexion.commit()

input("\nSe cargo tabla 'director' en 'movies.db'.")

for i in lista_idioma:
    inserta="INSERT INTO idioma values (?, ?);"
    cur.execute(inserta, i)
    conexion.commit()

input("\nSe cargo tabla 'idioma' en 'movies.db'.")

for i in lista_tematica:
    inserta="INSERT INTO tematica values (?, ?);"
    cur.execute(inserta, i)
    conexion.commit()

input("\nSe cargo tabla 'tematica' en 'movies.db'.")

for i in lista_actores:
    inserta="INSERT INTO actores values (?, ?, ?, ?);"
    cur.execute(inserta, i)
    conexion.commit()

input("\nSe cargo tabla 'actores' en 'movies.db'.")

for i in lista_peliculas:
    inserta="INSERT INTO peliculas values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?);"
    cur.execute(inserta, i)
    conexion.commit()

input("\nSe cargo tabla 'peliculas' en 'movies.db'.")


conexion.close()
